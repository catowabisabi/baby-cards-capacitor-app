from __future__ import annotations

import argparse
import copy
import json
import random
import shutil
import subprocess
import tempfile
import time
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from PIL import Image


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_WORKBOOK = BASE_DIR / "BabyCards_900卡詞彙清單.xlsx"
DEFAULT_WORKFLOW = BASE_DIR / "Z Image - Logo.json"
DEFAULT_OUTPUT_ROOT = PROJECT_DIR / "comfyui-output"
CAREER_PROMPT_DETAILS_PATH = BASE_DIR / "career-prompt-details.json"
CARS_PROMPT_DETAILS_PATH = BASE_DIR / "cars-prompt-details.json"

TOPIC_SHEETS = [
    "animals",
    "colors",
    "cars",
    "fruits",
    "food",
    "family",
    "body",
    "clothes",
    "toys",
    "dinosaurs",
    "career",
]

PROMPT_TEMPLATES = {
    "animals": (
        "professional realistic studio photograph of {english}, one single animal, "
        "full body, whole body clearly visible, slight three-quarter side view, "
        "body angled a little so the animal silhouette is easy for young children "
        "to recognize, centered, plain pure white seamless background filling the "
        "entire frame, soft even studio lighting, sharp focus, high detail, no text, "
        "no watermark, no border, no frame"
    ),
    "colors": (
        "professional realistic studio photograph representing {english}, one single "
        "simple object or swatch, centered, plain pure white seamless background "
        "filling the entire frame, soft even studio lighting, sharp focus, high detail, "
        "no text, no watermark, no border, no frame"
    ),
    "cars": (
        "professional realistic studio photograph of {english}, one single transportation "
        "vehicle, full body, whole vehicle clearly visible, slight three-quarter side view, "
        "centered, plain pure white seamless background filling the entire frame, soft even "
        "studio lighting, sharp focus, high detail, no text, no watermark, no border, no frame"
    ),
    "vehicles": (
        "professional realistic studio photograph of {english}, one single transportation "
        "vehicle, full body, whole vehicle clearly visible, slight three-quarter side view, "
        "centered, plain pure white seamless background filling the entire frame, soft even "
        "studio lighting, sharp focus, high detail, no text, no watermark, no border, no frame"
    ),
    "transportation": (
        "professional realistic studio photograph of {english}, one single transportation "
        "vehicle, full body, whole vehicle clearly visible, slight three-quarter side view, "
        "centered, plain pure white seamless background filling the entire frame, soft even "
        "studio lighting, sharp focus, high detail, no text, no watermark, no border, no frame"
    ),
    "fruits": (
        "professional realistic studio photograph of {english}, one whole fruit placed "
        "beside one cut-open half or clean slice of the same fruit showing the inside, "
        "only two pieces total, centered, plain pure white seamless background filling "
        "the entire frame, soft even studio lighting, sharp focus, high detail, no text, "
        "no watermark, no border, no frame"
    ),
    "food": (
        "professional realistic studio photograph of {english}, one single food item "
        "or simple serving, centered, plain pure white seamless background filling "
        "the entire frame, soft even studio lighting, sharp focus, high detail, no text, "
        "no watermark, no border, no frame"
    ),
    "family": (
        "professional realistic studio photograph of {english}, one single person or "
        "character, full body, whole body clearly visible, centered, plain pure white "
        "seamless background filling the entire frame, soft even studio lighting, "
        "sharp focus, high detail, child friendly, no text, no watermark, no border, no frame"
    ),
    "body": (
        "professional realistic studio photograph of one adult person only, age 25 to 40, "
        "modest clothing if clothing is visible, "
        "showing {english}, centered, plain pure white seamless background filling the "
        "entire frame, soft even studio lighting, sharp focus, high detail, adult only, "
        "no child, no baby, no kid, no printed card, no paper, no sign, no label, no text, "
        "no letters, no illustration, no cartoon, no watermark, no border, no frame"
    ),
    "clothes": (
        "professional realistic studio photograph of {english}, one single clothing item "
        "or accessory, full item clearly visible, centered, plain pure white seamless "
        "background filling the entire frame, soft even studio lighting, sharp focus, "
        "high detail, no text, no watermark, no border, no frame"
    ),
    "toys": (
        "professional realistic studio photograph of {english}, one single toy, full item "
        "clearly visible, centered, plain pure white seamless background filling the entire "
        "frame, soft even studio lighting, sharp focus, high detail, child friendly, "
        "no text, no watermark, no border, no frame"
    ),
    "dinosaurs": (
        "professional realistic studio photograph of {english}, one single dinosaur, "
        "full body, whole body clearly visible, slight three-quarter side view, body "
        "angled a little so the dinosaur silhouette is easy for young children to recognize, "
        "centered, plain pure white seamless background filling the entire frame, soft even "
        "studio lighting, sharp focus, high detail, child friendly, no text, no watermark, "
        "no border, no frame"
    ),
    "career": (
        "professional realistic studio photograph of one {english} actively working, "
        "natural workplace scene, clear work action, recognizable professional clothing "
        "and simple work tools, child friendly, centered, soft even lighting, sharp focus, "
        "high detail"
    ),
    "jobs": (
        "professional realistic studio photograph of one {english} actively working, "
        "natural workplace scene, clear work action, recognizable professional clothing "
        "and simple work tools, child friendly, centered, soft even lighting, sharp focus, "
        "high detail"
    ),
    "occupations": (
        "professional realistic studio photograph of one {english} actively working, "
        "natural workplace scene, clear work action, recognizable professional clothing "
        "and simple work tools, child friendly, centered, soft even lighting, sharp focus, "
        "high detail"
    ),
}


@dataclass
class Card:
    sheet: str
    number: int
    english: str
    chinese: str
    file_id: str
    notes: str = ""


@dataclass
class CardPaths:
    output_dir: Path
    json_path: Path
    webp_path: Path


@dataclass
class CardStatus:
    complete: bool
    reason: str


@dataclass
class HardwareStatus:
    gpu_temp_c: int | None = None
    vram_used_mb: int | None = None
    vram_total_mb: int | None = None
    cpu_temp_c: float | None = None

    @property
    def vram_used_percent(self) -> float | None:
        if not self.vram_used_mb or not self.vram_total_mb:
            return None
        return self.vram_used_mb / self.vram_total_mb * 100


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read BabyCards workbook, create JSON metadata, and generate card images with ComfyUI."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--workflow", type=Path, default=DEFAULT_WORKFLOW)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--server", default="http://127.0.0.1:8188")
    parser.add_argument("--sheets", default=",".join(TOPIC_SHEETS), help="Comma-separated sheet names.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of cards to process.")
    parser.add_argument("--start-after", default=None, help="Skip cards until after this file id.")
    parser.add_argument("--overwrite", action="store_true", help="Regenerate existing .webp/.json files.")
    parser.add_argument("--dry-run", action="store_true", help="Print planned work without writing files or calling ComfyUI.")
    parser.add_argument("--status-only", action="store_true", help="Only scan existing files and print progress.")
    parser.add_argument("--timeout", type=int, default=900, help="Seconds to wait for each ComfyUI prompt.")
    parser.add_argument("--poll-interval", type=float, default=2.0)
    parser.add_argument("--sleep-after-card", type=float, default=2.0, help="Seconds to sleep after each generated card.")
    parser.add_argument("--max-gpu-temp", type=int, default=82, help="Wait before the next card when GPU temp is above this C. Use 0 to disable.")
    parser.add_argument("--max-vram-percent", type=int, default=92, help="Wait when VRAM use is above this percent. Use 0 to disable.")
    parser.add_argument("--max-cpu-temp", type=int, default=90, help="Wait when readable CPU temp is above this C. Use 0 to disable.")
    parser.add_argument("--cooldown-interval", type=float, default=10.0, help="Seconds between hardware cooldown checks.")
    parser.add_argument("--seed", type=int, default=None, help="Fixed seed. Default uses a random seed per card.")
    parser.add_argument("--width", type=int, default=1024)
    parser.add_argument("--height", type=int, default=1024)
    parser.add_argument("--final-size", type=int, default=512)
    return parser.parse_args()


def load_cards(workbook_path: Path, sheets: list[str]) -> list[Card]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    cards: list[Card] = []

    for sheet_name in sheets:
        if sheet_name not in workbook.sheetnames:
            print(f"WARN sheet not found, skipping: {sheet_name}")
            continue

        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            number, english, chinese, file_id, notes = row[1], row[2], row[3], row[4], row[8]
            if not english or not chinese or not file_id:
                continue
            cards.append(
                Card(
                    sheet=sheet_name,
                    number=int(number),
                    english=str(english).strip(),
                    chinese=str(chinese).strip(),
                    file_id=str(file_id).strip(),
                    notes=str(notes).strip() if notes else "",
                )
            )

    if not cards:
        raise ValueError(f"No cards loaded from workbook: {workbook_path}")

    return cards


def make_prompt(card: Card) -> str:
    if card.sheet == "body":
        return make_body_prompt(card)
    if card.sheet in {"cars", "vehicles", "transportation"}:
        return make_cars_prompt(card)
    if card.sheet in {"career", "jobs", "occupations"}:
        return make_career_prompt(card)
    template = PROMPT_TEMPLATES.get(card.sheet, PROMPT_TEMPLATES["food"])
    prompt = template.format(english=card.english)
    if card.notes:
        prompt = f"{prompt}, specific direction: {card.notes}"
    return prompt


def load_career_prompt_details() -> dict[str, str]:
    if not CAREER_PROMPT_DETAILS_PATH.exists():
        return {}
    return json.loads(CAREER_PROMPT_DETAILS_PATH.read_text(encoding="utf-8"))


def make_career_prompt(card: Card) -> str:
    details = load_career_prompt_details()
    action = details.get(card.file_id) or f"one adult {card.english} doing a clear, easy to understand work task"
    return (
        "professional realistic studio photograph, child friendly, natural workplace scene, "
        f"{action}, clear simple action that a young child can understand, centered, "
        "soft even lighting, sharp focus, high detail"
    )


def load_cars_prompt_details() -> dict[str, str]:
    if not CARS_PROMPT_DETAILS_PATH.exists():
        return {}
    return json.loads(CARS_PROMPT_DETAILS_PATH.read_text(encoding="utf-8"))


def make_cars_prompt(card: Card) -> str:
    details = load_cars_prompt_details()
    vehicle = details.get(card.file_id) or f"{card.english}, clearly recognizable vehicle shape"
    return (
        "professional realistic studio photograph of one single transportation vehicle, "
        f"{vehicle}, full vehicle clearly visible, slight three-quarter side view, "
        "distinctive iconic features easy for young children to recognize, centered, "
        "plain pure white seamless background filling the entire frame, soft even studio "
        "lighting, sharp focus, high detail, no text, no watermark, no border, no frame"
    )


def make_body_prompt(card: Card) -> str:
    base = (
        "professional realistic studio photograph, plain pure white seamless background "
        "filling the entire frame, soft even studio lighting, sharp focus, high detail, "
        "real human adult body part or adult person only, age 25 to 40, centered, "
        "no child, no baby, no kid, no teenager, no printed card, no paper, no poster, "
        "no sign, no label, no text, no letters, no illustration, no cartoon, no drawing, "
        "no watermark, no border, no frame"
    )
    body_parts = {
        "head": "isolated close-up of one adult head only, neutral expression",
        "body": "isolated adult torso only, modest plain neutral shirt, no full body",
        "hand": "isolated close-up of one adult hand only, palm and fingers clearly visible",
        "foot": "isolated close-up of one adult foot only, clean neutral pose",
        "knee": "isolated close-up of one adult knee only, modest neutral clothing if visible",
        "toe": "isolated close-up of adult toes only, toes clearly visible",
        "finger": "isolated macro close-up of one adult index finger only, the finger fills most of the frame",
        "thumb": "isolated close-up of one adult thumb only, thumb clearly visible",
        "eye": "isolated close-up of one adult eye only",
        "ear": "isolated close-up of one adult ear only",
        "mouth": "isolated close-up of one adult mouth only, neutral expression",
        "nose": "isolated close-up of one adult nose only",
        "chin": "isolated close-up of one adult chin only",
        "cheek": "isolated close-up of one adult cheek only",
        "shoulder": "isolated close-up of one adult shoulder only, modest neutral clothing if visible",
        "back": "isolated close-up of one adult back only, modest neutral clothing if visible",
        "hair": "isolated close-up of adult hair only",
        "tongue": "isolated close-up of one adult mouth with tongue visible, neutral medical-style pose",
        "teeth": "isolated close-up of one adult mouth with teeth clearly visible, gentle natural smile",
    }
    emotions = {
        "happy": "adult face close-up only, natural happy smile",
        "sad": "adult face close-up only, natural sad expression",
        "crying": "adult face close-up only, gentle crying expression with a few tears",
        "sleepy": "adult face close-up only, sleepy expression",
        "tired": "adult face close-up only, tired expression",
    }
    action_details = {
        "running": "one adult running, full body clearly visible",
        "walking": "one adult walking, full body clearly visible",
        "jumping": "one adult jumping, full body clearly visible",
        "sitting-on-a-chair": "one adult sitting on a simple chair",
        "standing-still": "one adult standing still, full body clearly visible",
        "waving-a-hand": "one adult waving one hand",
        "clapping-hands": "one adult clapping hands",
        "pointing-at-a-star": "one adult pointing at a simple yellow star prop",
        "catching-a-soft-ball": "close-up of one adult hand catching a soft ball",
        "throwing-a-soft-ball": "one adult throwing a soft ball",
        "kicking-a-ball": "one adult kicking a ball",
        "climbing-a-tree": "one adult climbing a small safe tree, full body clearly visible",
        "reading-a-book": "one adult reading a book",
        "drinking-water": "one adult drinking water from a plain cup",
        "eating-an-apple": "one adult eating an apple",
        "brushing-teeth": "one adult brushing teeth with a toothbrush",
        "washing-hands": "one adult washing hands",
        "holding-a-cup": "one adult holding a plain cup",
        "hugging-a-teddy-bear": "one adult hugging a teddy bear",
        "sleeping-in-bed": "one adult sleeping in a simple bed, modest and fully covered",
    }

    detail = body_parts.get(card.file_id) or emotions.get(card.file_id) or action_details.get(card.file_id)
    if not detail:
        detail = f"one adult person showing {card.english}"
    return f"{base}, subject: {detail}"


def expected_metadata(card: Card) -> dict[str, str]:
    return {"en": card.english, "cn": card.chinese}


def paths_for_card(output_root: Path, card: Card) -> CardPaths:
    output_dir = output_root / card.sheet
    return CardPaths(
        output_dir=output_dir,
        json_path=output_dir / f"{card.file_id}.json",
        webp_path=output_dir / f"{card.file_id}.webp",
    )


def inspect_card_status(card: Card, paths: CardPaths, final_size: int) -> CardStatus:
    if not paths.json_path.exists() and not paths.webp_path.exists():
        return CardStatus(False, "missing json+webp")
    if not paths.json_path.exists():
        return CardStatus(False, "missing json")
    if not paths.webp_path.exists():
        return CardStatus(False, "missing webp")

    try:
        metadata = json.loads(paths.json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return CardStatus(False, f"invalid json: {exc}")

    if metadata != expected_metadata(card):
        return CardStatus(False, "json mismatch")

    try:
        with Image.open(paths.webp_path) as image:
            image.verify()
        with Image.open(paths.webp_path) as image:
            if image.format != "WEBP":
                return CardStatus(False, f"not webp: {image.format}")
            if image.size != (final_size, final_size):
                return CardStatus(False, f"wrong size: {image.size[0]}x{image.size[1]}")
    except Exception as exc:
        return CardStatus(False, f"invalid webp: {exc}")

    return CardStatus(True, "complete")


def summarize_status(cards: list[Card], output_root: Path, final_size: int) -> tuple[int, dict[str, int]]:
    counts: dict[str, int] = {}
    complete = 0
    for card in cards:
        status = inspect_card_status(card, paths_for_card(output_root, card), final_size)
        counts[status.reason] = counts.get(status.reason, 0) + 1
        if status.complete:
            complete += 1
    return complete, counts


def read_gpu_status() -> HardwareStatus:
    nvidia_smi = shutil.which("nvidia-smi")
    if not nvidia_smi:
        return HardwareStatus()

    command = [
        nvidia_smi,
        "--query-gpu=temperature.gpu,memory.used,memory.total",
        "--format=csv,noheader,nounits",
    ]
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=10, check=True)
        first_gpu = result.stdout.strip().splitlines()[0]
        temp, used, total = [part.strip() for part in first_gpu.split(",")[:3]]
        return HardwareStatus(gpu_temp_c=int(temp), vram_used_mb=int(used), vram_total_mb=int(total))
    except Exception:
        return HardwareStatus()


def read_cpu_temp_c() -> float | None:
    wmic = shutil.which("wmic")
    if not wmic:
        return None

    command = [
        wmic,
        "/namespace:\\\\root\\wmi",
        "PATH",
        "MSAcpi_ThermalZoneTemperature",
        "get",
        "CurrentTemperature",
        "/value",
    ]
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=10, check=True)
        temps: list[float] = []
        for line in result.stdout.splitlines():
            line = line.strip()
            if not line.startswith("CurrentTemperature="):
                continue
            raw = float(line.split("=", 1)[1])
            # WMI reports tenths of Kelvin.
            temps.append(raw / 10 - 273.15)
        return max(temps) if temps else None
    except Exception:
        return None


def read_hardware_status(include_cpu: bool) -> HardwareStatus:
    status = read_gpu_status()
    if include_cpu:
        status.cpu_temp_c = read_cpu_temp_c()
    return status


def format_hardware_status(status: HardwareStatus) -> str:
    parts: list[str] = []
    if status.gpu_temp_c is not None:
        parts.append(f"gpu={status.gpu_temp_c}C")
    if status.vram_used_mb is not None and status.vram_total_mb is not None:
        percent = status.vram_used_percent
        parts.append(f"vram={status.vram_used_mb}/{status.vram_total_mb}MB ({percent:.0f}%)")
    if status.cpu_temp_c is not None:
        parts.append(f"cpu={status.cpu_temp_c:.0f}C")
    return ", ".join(parts) if parts else "hardware sensors unavailable"


def wait_for_hardware_cooldown(args: argparse.Namespace) -> None:
    include_cpu = args.max_cpu_temp > 0
    warned_no_gpu = False
    warned_no_cpu = False

    while True:
        status = read_hardware_status(include_cpu)
        reasons: list[str] = []

        if args.max_gpu_temp > 0:
            if status.gpu_temp_c is None:
                if not warned_no_gpu:
                    print("WARN gpu temp unavailable; install/enable nvidia-smi to check GPU heat")
                    warned_no_gpu = True
            elif status.gpu_temp_c >= args.max_gpu_temp:
                reasons.append(f"GPU {status.gpu_temp_c}C >= {args.max_gpu_temp}C")

        vram_percent = status.vram_used_percent
        if args.max_vram_percent > 0 and vram_percent is not None and vram_percent >= args.max_vram_percent:
            reasons.append(f"VRAM {vram_percent:.0f}% >= {args.max_vram_percent}%")

        if args.max_cpu_temp > 0:
            if status.cpu_temp_c is None:
                if not warned_no_cpu:
                    print("WARN cpu temp unavailable on this system; CPU heat check skipped")
                    warned_no_cpu = True
            elif status.cpu_temp_c >= args.max_cpu_temp:
                reasons.append(f"CPU {status.cpu_temp_c:.0f}C >= {args.max_cpu_temp}C")

        if not reasons:
            print(f"HW OK {format_hardware_status(status)}")
            return

        print(
            f"COOLDOWN {'; '.join(reasons)}; {format_hardware_status(status)}; "
            f"sleeping {args.cooldown_interval:g}s"
        )
        time.sleep(args.cooldown_interval)


def request_json(url: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(request) as response:
        return json.loads(response.read().decode("utf-8"))


def queue_prompt(server: str, workflow: dict[str, Any], client_id: str) -> str:
    response = request_json(f"{server.rstrip('/')}/prompt", {"prompt": workflow, "client_id": client_id})
    return response["prompt_id"]


def wait_for_outputs(server: str, prompt_id: str, timeout: int, poll_interval: float) -> list[dict[str, Any]]:
    deadline = time.time() + timeout
    history_url = f"{server.rstrip('/')}/history/{prompt_id}"

    while time.time() < deadline:
        history = request_json(history_url)
        item = history.get(prompt_id)
        if item and item.get("outputs"):
            images: list[dict[str, Any]] = []
            for output in item["outputs"].values():
                images.extend(output.get("images", []))
            if images:
                return images
        time.sleep(poll_interval)

    raise TimeoutError(f"Timed out waiting for ComfyUI prompt: {prompt_id}")


def download_image(server: str, image_info: dict[str, Any], destination: Path, final_size: int) -> None:
    query = urllib.parse.urlencode(
        {
            "filename": image_info["filename"],
            "subfolder": image_info.get("subfolder", ""),
            "type": image_info.get("type", "output"),
        }
    )
    image_url = f"{server.rstrip('/')}/view?{query}"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".source", dir=destination.parent) as temp_file:
        temp_path = Path(temp_file.name)
        with urllib.request.urlopen(image_url) as response:
            temp_file.write(response.read())

    with Image.open(temp_path) as image:
        image = image.convert("RGBA")
        if image.size != (final_size, final_size):
            image.thumbnail((final_size, final_size), Image.Resampling.LANCZOS)
            canvas = Image.new("RGBA", (final_size, final_size), (255, 255, 255, 0))
            x = (final_size - image.width) // 2
            y = (final_size - image.height) // 2
            canvas.alpha_composite(image, (x, y))
            image = canvas
        tmp_destination = destination.with_suffix(".tmp.webp")
        image.save(tmp_destination, "WEBP", quality=95, method=6)
        tmp_destination.replace(destination)

    temp_path.unlink(missing_ok=True)


def write_metadata(destination: Path, card: Card) -> None:
    tmp_destination = destination.with_suffix(".tmp.json")
    tmp_destination.write_text(json.dumps(expected_metadata(card), ensure_ascii=False), encoding="utf-8")
    tmp_destination.replace(destination)


def build_workflow(
    template: dict[str, Any],
    card: Card,
    filename_prefix: str,
    seed: int | None,
    width: int,
    height: int,
) -> dict[str, Any]:
    workflow = copy.deepcopy(template)
    workflow["45"]["inputs"]["text"] = make_prompt(card)
    workflow["9"]["inputs"]["filename_prefix"] = filename_prefix
    workflow["41"]["inputs"]["width"] = width
    workflow["41"]["inputs"]["height"] = height
    workflow["44"]["inputs"]["seed"] = seed if seed is not None else random.randint(1, 2**63 - 1)
    return workflow


def main() -> None:
    args = parse_args()
    sheets = [sheet.strip() for sheet in args.sheets.split(",") if sheet.strip()]
    cards = load_cards(args.workbook, sheets)

    if args.start_after:
        ids = [card.file_id for card in cards]
        if args.start_after not in ids:
            raise ValueError(f"--start-after id not found: {args.start_after}")
        cards = cards[ids.index(args.start_after) + 1 :]

    if args.limit is not None:
        cards = cards[: args.limit]

    workflow_template = json.loads(args.workflow.read_text(encoding="utf-8"))
    client_id = str(uuid.uuid4())
    processed = 0
    skipped = 0
    failed = 0

    complete, status_counts = summarize_status(cards, args.output_root, args.final_size)
    remaining = len(cards) - complete
    print(f"STATUS total_selected={len(cards)} complete={complete} remaining={remaining}")
    for reason, count in sorted(status_counts.items()):
        print(f"STATUS {reason}: {count}")

    if args.status_only:
        print("DONE status-only")
        return

    for card in cards:
        paths = paths_for_card(args.output_root, card)
        status = inspect_card_status(card, paths, args.final_size)

        if not args.overwrite and status.complete:
            skipped += 1
            print(f"SKIP existing {card.sheet}/{card.file_id}")
            continue

        if not args.overwrite and status.reason != "missing json+webp":
            print(f"REPAIR {card.sheet}/{card.file_id}: {status.reason}")

        paths.output_dir.mkdir(parents=True, exist_ok=True)
        prompt = make_prompt(card)
        print(f"PROCESS {card.sheet}/{card.file_id}: {card.english} -> {paths.webp_path}")
        print(f"PROMPT {prompt}")

        if args.dry_run:
            processed += 1
            continue

        wait_for_hardware_cooldown(args)
        write_metadata(paths.json_path, card)
        filename_prefix = f"baby-cards-tmp/{card.sheet}/{card.file_id}"
        workflow = build_workflow(
            workflow_template,
            card,
            filename_prefix=filename_prefix,
            seed=args.seed,
            width=args.width,
            height=args.height,
        )
        try:
            prompt_id = queue_prompt(args.server, workflow, client_id)
            images = wait_for_outputs(args.server, prompt_id, args.timeout, args.poll_interval)
            download_image(args.server, images[-1], paths.webp_path, args.final_size)
            final_status = inspect_card_status(card, paths, args.final_size)
            if not final_status.complete:
                raise RuntimeError(f"output verification failed: {final_status.reason}")
            processed += 1
            if args.sleep_after_card > 0:
                print(f"SLEEP {args.sleep_after_card:g}s")
                time.sleep(args.sleep_after_card)
        except Exception as exc:
            failed += 1
            print(f"ERROR {card.sheet}/{card.file_id}: {exc}")
            raise

    complete, _ = summarize_status(cards, args.output_root, args.final_size)
    print(
        f"DONE processed={processed} skipped={skipped} failed={failed} "
        f"complete={complete} remaining={len(cards) - complete} total_selected={len(cards)}"
    )


if __name__ == "__main__":
    main()
